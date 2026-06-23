#!/usr/bin/env python3
"""Extract data from [Junho] Carteira CS spreadsheet into JSON for dashboard."""
import json, sys, warnings
from datetime import datetime
import openpyxl
import pandas as pd

warnings.filterwarnings("ignore")

XLSX = sys.argv[1] if len(sys.argv) > 1 else "/tmp/new_sheet.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "docs/data.json"


def safe(v, default=""):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return default
    return v


def to_float(v, default=0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def to_int(v, default=0):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def extract_meta_geral(wb):
    ws = wb["Meta Geral 062025"]
    rows = {}
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        rows[r] = vals

    return {
        "contratos": {
            "meta_geral": to_int(rows[6][1]),
            "realizado_geral": to_int(rows[6][3]),
            "pct_geral": round(to_float(rows[6][5]) * 100, 1),
        },
        "retencao": {
            "clientes_totais": to_int(rows[6][7]),
            "pct_retencao": round(to_float(rows[6][9]) * 100, 1),
            "clientes_reverter": round(to_float(rows[6][11]), 1),
        },
        "onboarding": {
            "meta": to_int(rows[10][1]),
            "realizado": to_int(rows[10][3]),
            "pct": round(to_float(rows[10][5]) * 100, 1),
            "clientes_revertidos": to_int(rows[10][7]),
            "pct_revertidos": round(to_float(rows[10][9]) * 100, 1),
            "atingimento_100": round(to_float(rows[10][11]) * 100, 1),
        },
        "ongoing": {
            "meta": to_int(rows[14][1]),
            "realizado": to_int(rows[14][3]),
            "pct": round(to_float(rows[14][5]) * 100, 1),
        },
    }


def extract_meta_individual(wb):
    ws = wb["Meta Individual 06-2026"]
    rows = {}
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        rows[r] = vals

    specialists = []

    # Camila (Onboarding) - rows 3-6
    specialists.append({
        "nome": "Camila",
        "tipo": "Onboarding",
        "contratos": {
            "clientes_carteira": to_int(rows[6][4]),
            "meta": to_int(rows[6][5]),
            "ativos": to_int(rows[6][6]),
            "pct": round(to_float(rows[6][7]) * 100, 1),
        },
        "segundo_contrato": {
            "clientes_carteira": to_int(rows[6][9]),
            "meta": round(to_float(rows[6][10]) * 100, 1),
            "realizado": to_int(rows[6][11]),
            "pct": round(to_float(rows[6][12]) * 100, 1),
        },
        "resultado_final": round(to_float(rows[4][14]) * 100, 1),
    })

    # Andressa (Ongoing) - rows 10-13
    specialists.append({
        "nome": "Andressa",
        "tipo": "Ongoing",
        "contratos": {
            "clientes_carteira": to_int(rows[13][4]),
            "meta": to_int(rows[13][5]),
            "ativos": to_int(rows[13][6]),
            "pct": round(to_float(rows[13][7]) * 100, 1),
        },
        "retencao": {
            "clientes_risco": to_int(rows[13][9]),
            "meta_reversao": round(to_float(rows[13][10]), 1),
            "revertidos": to_int(rows[13][11]),
            "pct": round(to_float(rows[13][12]) * 100, 1),
        },
        "resultado_final": round(to_float(rows[11][14]) * 100, 1),
    })

    # Emanuelle (Ongoing) - rows 17-20
    specialists.append({
        "nome": "Emanuelle",
        "tipo": "Ongoing",
        "contratos": {
            "clientes_carteira": to_int(rows[20][4]),
            "meta": to_int(rows[20][5]),
            "ativos": to_int(rows[20][6]),
            "pct": round(to_float(rows[20][7]) * 100, 1),
        },
        "retencao": {
            "clientes_risco": to_int(rows[20][9]),
            "meta_reversao": round(to_float(rows[20][10]), 1),
            "revertidos": to_int(rows[20][11]),
            "pct": round(to_float(rows[20][12]) * 100, 1),
        },
        "resultado_final": round(to_float(rows[18][14]) * 100, 1),
    })

    return specialists


def extract_visao_geral(wb):
    ws = wb["Visão Geral da Carteira"]
    summary = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    clients = []
    status_dist = {}
    carteira_dist = {}
    lost_dist = {}

    for r in range(3, ws.max_row + 1):
        corretor = ws.cell(r, 2).value
        if not corretor:
            continue
        status = safe(ws.cell(r, 6).value, "Desconhecido")
        carteira = safe(ws.cell(r, 7).value, "Desconhecido")
        lost = safe(ws.cell(r, 10).value, "—")
        dias = to_float(ws.cell(r, 5).value, None)

        clients.append({
            "corretor": str(corretor),
            "id_cliente": safe(ws.cell(r, 3).value),
            "data_primeiro_contrato": str(safe(ws.cell(r, 4).value))[:10],
            "dias_sem_contrato": dias,
            "status": status,
            "carteira": carteira,
            "contratos_ate_abril": to_int(ws.cell(r, 8).value),
            "contrato_ativo_junho": to_int(ws.cell(r, 9).value),
            "lost": str(lost),
        })

        status_dist[status] = status_dist.get(status, 0) + 1
        carteira_dist[carteira] = carteira_dist.get(carteira, 0) + 1
        lost_dist[str(lost)] = lost_dist.get(str(lost), 0) + 1

    return {
        "total_clientes": to_int(summary[2]),
        "media_contratos": round(to_float(summary[7]), 2),
        "contratos_emitidos_junho": to_int(summary[8]),
        "contratos_ativos": to_int(summary[9]),
        "status_distribution": status_dist,
        "carteira_distribution": carteira_dist,
        "lost_distribution": lost_dist,
        "clients": clients,
    }


def extract_churn(wb):
    ws = wb["Visão Churn"]
    churn = []
    for r in range(3, 14):
        label = ws.cell(r, 4).value
        count = to_int(ws.cell(r, 6).value)
        if label:
            churn.append({"range": str(label), "count": count})
    return churn


def extract_clientes_perdidos(wb):
    ws = wb["Clientes Perdidos"]
    clients = []
    motivos = {}
    for r in range(3, ws.max_row + 1):
        nome = ws.cell(r, 1).value
        if not nome:
            continue
        motivo = safe(ws.cell(r, 8).value, "Nenhum")
        clients.append({
            "corretor": str(nome),
            "id_cliente": safe(ws.cell(r, 2).value),
            "data_primeiro_contrato": str(safe(ws.cell(r, 3).value))[:10],
            "dias_sem_contrato": to_int(ws.cell(r, 4).value),
            "status": safe(ws.cell(r, 5).value),
            "contratos_anteriores": to_int(ws.cell(r, 6).value),
            "entrou_lost": str(safe(ws.cell(r, 7).value))[:10],
            "motivo": motivo,
            "consideracoes": safe(ws.cell(r, 9).value),
        })
        motivos[motivo] = motivos.get(motivo, 0) + 1

    return {
        "total": len(clients),
        "motivos": motivos,
        "clients": clients,
    }


def extract_foco_semana(wb):
    ws = wb["Foco da semana"]
    items = []
    for r in range(2, ws.max_row + 1):
        nome = ws.cell(r, 1).value
        if not nome:
            continue
        items.append({
            "nome": str(nome),
            "status": safe(ws.cell(r, 2).value),
            "especialista": safe(ws.cell(r, 3).value),
            "contratos_previstos": to_int(ws.cell(r, 4).value),
            "valor_2k": str(safe(ws.cell(r, 5).value)).lower() == "true",
            "atuacao_jamile": str(safe(ws.cell(r, 6).value)).lower() == "true",
            "telefone": safe(ws.cell(r, 7).value),
            "observacao": safe(ws.cell(r, 8).value),
        })
    return items


def extract_reunioes(wb):
    ws = wb["Reuniões Realizadas"]
    dates = []
    for c in range(3, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and "Link" not in str(v):
            dates.append(str(v)[:10])

    by_specialist = {}
    meetings = []
    for r in range(2, ws.max_row + 1):
        specialist = ws.cell(r, 1).value
        client = ws.cell(r, 2).value
        if not specialist or not client:
            continue
        for i, c in enumerate(range(3, 3 + len(dates))):
            if str(ws.cell(r, c).value).lower() == "true":
                meetings.append({
                    "especialista": str(specialist),
                    "cliente": str(client),
                    "data": dates[i] if i < len(dates) else "",
                })
                by_specialist[str(specialist)] = by_specialist.get(str(specialist), 0) + 1

    return {
        "dates": dates,
        "total": len(meetings),
        "by_specialist": by_specialist,
        "meetings": meetings,
    }


def extract_renovacoes(wb):
    ws = wb["Renovações de Contratos"]
    contracts = []
    for r in range(2, ws.max_row + 1):
        added = ws.cell(r, 1).value
        if not added:
            continue
        contracts.append({
            "adicionado": str(safe(ws.cell(r, 1).value))[:10],
            "planejamento_termino": str(safe(ws.cell(r, 2).value))[:10],
            "codigo": safe(ws.cell(r, 3).value),
            "responsavel": safe(ws.cell(r, 6).value),
            "email": safe(ws.cell(r, 7).value),
            "valor_aluguel": to_float(ws.cell(r, 8).value),
            "data_inicio": str(safe(ws.cell(r, 12).value))[:10],
            "data_prevista_termino": str(safe(ws.cell(r, 13).value))[:10],
            "data_real_termino": str(safe(ws.cell(r, 14).value))[:10],
        })
    return contracts


def extract_carteira_individual(wb, sheet_name, specialist_name):
    ws = wb[sheet_name]
    summary = [ws.cell(1, c).value for c in range(1, min(15, ws.max_column + 1))]
    header_row = 3
    clients = []
    for r in range(header_row + 1, ws.max_row + 1):
        nome = ws.cell(r, 3).value
        if not nome:
            continue
        clients.append({
            "corretor": str(nome),
            "dias_sem_contrato": to_int(ws.cell(r, 6).value),
            "status": safe(ws.cell(r, 7).value),
            "contratos_ate_abril": to_int(ws.cell(r, 8).value),
            "contratos_maio": to_float(ws.cell(r, 9).value),
            "contratos_junho": to_int(ws.cell(r, 10).value),
            "lost": safe(ws.cell(r, 11).value),
            "termometro": safe(ws.cell(r, 12).value),
            "reativado": str(safe(ws.cell(r, 13).value)).lower() == "true",
            "status_pipeline": safe(ws.cell(r, 14).value),
        })
    return {
        "total": len(clients),
        "clients": clients,
    }


def extract_backups(wb):
    ws = wb["Backups"]
    backups = []
    for r in range(1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        url = ws.cell(r, 2).value
        if name:
            backups.append({"mes": str(name), "url": str(safe(url))})
    return backups


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    data = {
        "titulo": "[Junho] Carteira CS",
        "periodo": "Junho 2025",
        "gerado_em": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "meta_geral": extract_meta_geral(wb),
        "meta_individual": extract_meta_individual(wb),
        "visao_geral": extract_visao_geral(wb),
        "churn": extract_churn(wb),
        "clientes_perdidos": extract_clientes_perdidos(wb),
        "foco_semana": extract_foco_semana(wb),
        "reunioes": extract_reunioes(wb),
        "renovacoes": extract_renovacoes(wb),
        "carteira_andressa": extract_carteira_individual(wb, "Carteira Andressa", "Andressa"),
        "carteira_emanuelle": extract_carteira_individual(wb, "Carteira Emanuelle", "Emanuelle"),
        "backups": extract_backups(wb),
    }

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    vg = data["visao_geral"]
    mg = data["meta_geral"]
    print(f"OK — {vg['total_clientes']} clients, "
          f"{mg['contratos']['realizado_geral']}/{mg['contratos']['meta_geral']} contracts, "
          f"{data['clientes_perdidos']['total']} lost, "
          f"{len(data['foco_semana'])} foco items, "
          f"{data['reunioes']['total']} meetings, "
          f"{len(data['renovacoes'])} renewals")


if __name__ == "__main__":
    main()
